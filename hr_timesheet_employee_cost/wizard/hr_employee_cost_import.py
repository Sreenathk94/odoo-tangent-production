# -*- coding: utf-8 -*-
from odoo import fields, models, api, _
from odoo.exceptions import UserError
import base64
from openpyxl import load_workbook
from io import BytesIO
import logging

_logger = logging.getLogger(__name__)


class HrEmployeeCostImportWizard(models.TransientModel):
    _name = 'hr.employee.cost.import.wizard'
    _description = 'Import Employee Costs Wizard'

    date_from = fields.Date('From Date', required=True)
    date_to = fields.Date('To Date', required=True)
    import_file = fields.Binary('Excel File', required=True)
    filename = fields.Char('Filename')

    def action_import(self):
        """Import employee costs from Excel"""

        # Validate file
        if not self.import_file:
            raise UserError(_('Please select a file to import.'))

        if self.date_from > self.date_to:
            raise UserError(_('From Date cannot be greater than To Date.'))

        # Validate file extension
        if not self.filename or not self.filename.lower().endswith(
            ('.xlsx', '.xlsm', '.xltx', '.xltm')
        ):
            raise UserError(_('Invalid file format. Only Excel files are supported.'))

        # Load Excel file
        try:
            file_data = base64.b64decode(self.import_file)
            wb = load_workbook(filename=BytesIO(file_data))
            ws = wb.active
        except Exception as e:
            raise UserError(_('Could not read Excel file: %s') % str(e))

        # Read header row
        headers = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                headers[col] = str(val).strip().lower()

        # Identify required columns
        first_col = next((c for c, h in headers.items() if 'first' in h or h == 'name'), None)
        last_col = next((c for c, h in headers.items() if 'last' in h or 'second' in h), None)
        cost_col = next((c for c, h in headers.items() if 'cost' in h), None)
        bio_col = next((c for c, h in headers.items() if 'bio' in h or 'biometric' in h), None)

        if not first_col or not cost_col:
            raise UserError(_('Excel must contain "First Name" and "Cost" columns.'))

        if not bio_col:
            raise UserError(_('Excel must contain "Biometric Code" column.'))

        # Create cost header
        cost_header = self.env['hr.employee.cost'].create({
            'name': _('Import %s to %s') % (self.date_from, self.date_to),
            'date_from': self.date_from,
            'date_to': self.date_to,
            'state': 'draft',
        })

        imported = 0
        skipped = []

        # Process rows
        for row in range(2, ws.max_row + 1):
            first = ws.cell(row=row, column=first_col).value
            last = ws.cell(row=row, column=last_col).value if last_col else ''
            cost = ws.cell(row=row, column=cost_col).value or 0.0
            bio_code = ws.cell(row=row, column=bio_col).value

            if not first:
                continue

            full_name = f"{first} {last}".strip()
            employee = self._find_employee_by_biocode(bio_code, full_name)

            if employee:
                self.env['hr.employee.cost.line'].create({
                    'cost_id': cost_header.id,
                    'employee_id': employee.id,
                    'cost': cost,
                })
                imported += 1
            else:
                skipped.append(full_name)
                _logger.warning(f"Employee '{full_name}' with bio_code '{bio_code}' not found.")

        # Confirm the header
        cost_header.action_confirm()

        # Build notification
        msg = _('%d employees imported successfully.') % imported
        if skipped:
            msg += _('\nSkipped employees: %s') % ', '.join(skipped[:10])
            if len(skipped) > 10:
                msg += _(' ...and %d more') % (len(skipped) - 10)

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import Completed'),
                'message': msg,
                'type': 'success' if imported else 'warning',
                'sticky': False,
            },
        }

    def _find_employee_by_biocode(self, bio_code, full_name):
        """Find ALL employees (active + archived) by bio_code and exact full name."""
        Employee = self.env['hr.employee'].with_context(active_test=False)

        bio = (str(bio_code) or '').strip()
        name = (str(full_name) or '').strip()

        if not bio and not name:
            return Employee.browse()  # empty recordset

        domain = []
        if bio:
            domain.append(('bio_code', '=', bio))
        if name:
            # exact (case-insensitive) full name match
            domain.append(('name', '=ilike', name))

        employees = Employee.search(domain)

        return employees

